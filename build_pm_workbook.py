#!/usr/bin/env python3

from __future__ import annotations

import argparse
import re
from collections import defaultdict
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


WORKDIR = Path(__file__).resolve().parent
DEFAULT_OUTPUT_DIR = WORKDIR / "outputs"
HEADER_FILL = PatternFill("solid", fgColor="F3F4F6")
TITLE_FILL = PatternFill("solid", fgColor="FFFFFF")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
BLACK_SIDE = Side(style="thin", color="000000")
GRAY_SIDE = Side(style="thin", color="D9D9D9")
HEADER_BORDER = Border(left=BLACK_SIDE, right=BLACK_SIDE, top=BLACK_SIDE, bottom=BLACK_SIDE)
BODY_BORDER = Border(left=GRAY_SIDE, right=GRAY_SIDE, top=GRAY_SIDE, bottom=GRAY_SIDE)
TITLE_FONT = Font(name="Calibri", size=12, bold=True, color="000000")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="000000")
BODY_FONT = Font(name="Calibri", size=11, color="000000")
HEADERS = [
    "Side",
    "Pos. Size Unit",
    "USD Exposure%",
    "USD Exposure",
    "Position size(Equity)",
    "Mark Price",
    "Latest Price",
    "Valuation Source",
    "Account Type",
    "Equity Valuation",
    "BS Delta",
    "PA Delta",
    "BS Gamma",
    "PA Gamma",
    "Vega",
    "Theta",
    "IV",
    "ATM Vol",
    "Exercise Prob.",
    "Model Price",
    "Volatility Source",
    "Account",
    "Protocol",
    "Pool",
    "Val. Ccy.",
    "Cash Balance",
    "UPnL (Settlement Ccy.)",
]
GROUP_HEADERS = {
    2: "Pos. Size Info",
    5: "Pos. Size Info",
    6: "Price Info",
    9: "Account Info",
    11: "Greeks Info",
    22: "Account Info",
}
COLUMN_WIDTHS = [
    8,
    15,
    15,
    15,
    20,
    13,
    14,
    17,
    23,
    17,
    13,
    12,
    12,
    12,
    9,
    9,
    9,
    11,
    15,
    13,
    17,
    24,
    12,
    12,
    11,
    15,
    23,
]

# Output order matches export.xlsx after its Symbol column is removed.
COLUMN_MAP = [
    ("side_str", False),
    ("position_size_unit_str", False),
    ("usd_exposure_ratio", True),
    ("usd_exposure", True),
    ("position_size_equity", True),
    ("mark_price", True),
    ("last_price", True),
    ("valuation_source_str", False),
    ("account_type_str", False),
    ("equity_valuation", True),
    ("bs_delta", True),
    ("delta", True),
    ("bs_gamma", True),
    ("gamma", True),
    ("vega", True),
    ("theta", True),
    ("iv_str", False),
    ("atm_iv_str", False),
    ("exercise_prob_str", False),
    ("model_price_str", False),
    ("volatility_source_str", False),
    ("account_alias", False),
    ("protocol_str", False),
    ("pool_str", False),
    ("requested_fund_valuation_currency", False),
    ("cash_balance", True),
    ("upnl_settlement_currency", True),
]


def numeric_value(value: Any) -> float | str:
    if value in (None, "", "-"):
        return "-"
    try:
        return float(Decimal(str(value)))
    except (InvalidOperation, ValueError):
        return str(value)


def exposure_value(row: dict[str, Any]) -> Decimal:
    try:
        return Decimal(str(row.get("usd_exposure") or 0))
    except (InvalidOperation, ValueError):
        return Decimal(0)


def display_value(row: dict[str, Any], field: str, numeric: bool) -> Any:
    value = row.get(field)
    if field == "requested_fund_valuation_currency" and value not in (None, "", "-"):
        return str(value).upper()
    if numeric:
        return numeric_value(value)
    return "-" if value in (None, "") else value


def selected_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    ordered = sorted(rows, key=exposure_value, reverse=True)
    if len(ordered) <= 20:
        return ordered
    return ordered[:10] + ordered[-10:]


def safe_sheet_name(name: str, used_names: set[str]) -> str:
    base = re.sub(r"[\[\]:*?/\\]", "-", name).strip() or "Portfolio"
    candidate = base[:31]
    suffix_number = 2
    while candidate.casefold() in used_names:
        suffix = f"~{suffix_number}"
        candidate = f"{base[:31 - len(suffix)]}{suffix}"
        suffix_number += 1
    used_names.add(candidate.casefold())
    return candidate


def create_layout(output_sheet, manager: str) -> None:
    output_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    output_sheet["A1"] = manager
    for column, heading in GROUP_HEADERS.items():
        output_sheet.cell(2, column).value = heading
    for column, heading in enumerate(HEADERS, start=1):
        output_sheet.cell(3, column).value = heading
    for column, width in enumerate(COLUMN_WIDTHS, start=1):
        output_sheet.column_dimensions[get_column_letter(column)].width = width
    output_sheet.row_dimensions[1].height = 27
    output_sheet.row_dimensions[2].height = 22
    output_sheet.row_dimensions[3].height = 25
    output_sheet.freeze_panes = "A4"


def add_rows(output_sheet, rows: list[dict[str, Any]]) -> None:
    for output_row, row in enumerate(rows, start=4):
        output_sheet.row_dimensions[output_row].height = 20
        for output_column, (field, numeric) in enumerate(COLUMN_MAP, start=1):
            output_cell = output_sheet.cell(output_row, output_column)
            output_cell.value = display_value(row, field, numeric)

    for row_number in range(4, output_sheet.max_row + 1):
        output_sheet.cell(row_number, 3).number_format = "0.00%"
        output_sheet.cell(row_number, 4).number_format = "#,##0.00"
        output_sheet.cell(row_number, 7).number_format = "#,##0.0000"
        for column in (5, 6, 10, 11, 12, 13, 14, 15, 16, 26, 27):
            output_sheet.cell(row_number, column).number_format = "#,##0.########"


def apply_clean_black_style(sheet) -> None:
    for row in sheet.iter_rows():
        for cell in row:
            cell.fill = WHITE_FILL
            cell.font = BODY_FONT
            cell.border = BODY_BORDER
            cell.alignment = Alignment(
                horizontal="right" if isinstance(cell.value, (int, float)) else "left",
                vertical="center",
                wrap_text=False,
            )

    for cell in sheet[1]:
        cell.fill = TITLE_FILL
        cell.font = TITLE_FONT
        cell.border = Border(bottom=BLACK_SIDE)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_number in (2, 3):
        for cell in sheet[row_number]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = HEADER_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    sheet.sheet_view.showGridLines = False


def build_pm_workbook(
    rows: Iterable[dict[str, Any]],
    *,
    output_path: Path,
) -> dict[str, int]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        manager = str(row.get("requested_fund_alias") or row.get("fund_alias") or "").strip()
        if manager:
            grouped[manager].append(row)

    if not grouped:
        raise ValueError("No portfolio manager rows were found.")

    output_workbook = Workbook()
    output_workbook.remove(output_workbook.active)
    used_names: set[str] = set()
    counts: dict[str, int] = {}

    for manager in sorted(grouped, key=str.casefold):
        manager_rows = selected_rows(grouped[manager])
        sheet = output_workbook.create_sheet(safe_sheet_name(manager, used_names))
        create_layout(sheet, manager)
        add_rows(sheet, manager_rows)
        apply_clean_black_style(sheet)
        counts[manager] = len(manager_rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_workbook.save(output_path)
    return counts


def read_raw_rows(source_path: Path) -> Iterable[dict[str, Any]]:
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    sheet = workbook["Raw Data"]
    rows = sheet.iter_rows(values_only=True)
    headers = list(next(rows))

    required_fields = {
        "requested_fund_alias",
        "fund_alias",
        *(field for field, _ in COLUMN_MAP),
    }
    indexes = {
        field: headers.index(field)
        for field in required_fields
        if field in headers
    }

    for values in rows:
        yield {field: values[index] for field, index in indexes.items()}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Create one formatted top/bottom exposure sheet per portfolio manager."
    )
    parser.add_argument("source", type=Path, help="API workbook containing the Raw Data sheet.")
    parser.add_argument("--output", type=Path)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    source_path = args.source.expanduser().resolve()
    output_path = (
        args.output.expanduser().resolve()
        if args.output
        else DEFAULT_OUTPUT_DIR / f"{source_path.stem}_pm_top_bottom.xlsx"
    )

    counts = build_pm_workbook(
        read_raw_rows(source_path),
        output_path=output_path,
    )
    print(f"Created {len(counts)} portfolio manager sheets.")
    print(f"Saved to: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
