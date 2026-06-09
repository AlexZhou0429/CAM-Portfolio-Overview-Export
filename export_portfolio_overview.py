#!/usr/bin/env python3

from __future__ import annotations

import argparse
import base64
import getpass
import hashlib
import hmac
import json
import os
import sys
import time
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import Request, urlopen
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from build_pm_workbook import build_pm_workbook


DEFAULT_BASE_URL = "https://psg.1token.tech/api/v1"
DEFAULT_DEPARTMENT = ""
DEFAULT_TAG = "SP Core"
DEFAULT_TIMEZONE = "Asia/Shanghai"
WORKDIR = Path(__file__).resolve().parent
OUTPUT_DIR = WORKDIR / "outputs"
MAX_FUNDS_PER_REQUEST = 30
DEFAULT_USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/136.0.0.0 Safari/537.36"
)

SUMMARY_COLUMNS = [
    ("Portfolio", "requested_fund_alias"),
    ("Portfolio ID", "requested_fund_name"),
    ("Snapshot Time", "snapshot_time_str"),
    ("Side", "side_str"),
    ("Symbol", "symbol_str"),
    ("Instrument", "instrument_str"),
    ("Account", "account_alias"),
    ("Venue", "venue_str"),
    ("Account Type", "account_type_str"),
    ("Position Size", "position_size_equity"),
    ("Unit", "position_size_unit_str"),
    ("USD Exposure %", "usd_exposure_ratio"),
    ("USD Exposure", "usd_exposure"),
    ("Equity Valuation", "equity_valuation"),
    ("Equity Currency", "equity_valuation_currency"),
    ("Mark Price", "mark_price"),
    ("Last Price", "last_price"),
    ("UPnL USD", "upnl_usd"),
    ("Underlying", "underlying_str"),
    ("Valuation Source", "valuation_source_str"),
]


class CamApiError(RuntimeError):
    pass


def decimal_or_zero(value: Any) -> Decimal:
    try:
        if value in (None, "", "-"):
            return Decimal(0)
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return Decimal(0)


def excel_value(value: Any) -> Any:
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, separators=(",", ":"))
    if value is None:
        return None
    return value


class CamClient:
    def __init__(
        self,
        *,
        base_url: str,
        department: str,
        api_key: str | None = None,
        api_secret: str | None = None,
        token: str | None = None,
        token_header: str = "api-token",
        timeout: int = 60,
    ) -> None:
        self.base_url = base_url.rstrip("/")
        self.department = department
        self.api_key = api_key
        self.api_secret = api_secret
        self.token = token
        self.token_header = token_header
        self.timeout = timeout

    def authentication_headers(
        self,
        method: str,
        request_path: str,
        body_text: str,
    ) -> dict[str, str]:
        if self.api_key and self.api_secret:
            timestamp = int(time.time())
            message = f"{method}{request_path}{timestamp}{body_text}".encode("utf-8")
            try:
                secret_bytes = base64.b64decode(self.api_secret)
            except (ValueError, TypeError) as exc:
                raise CamApiError("The CAM API Secret is not valid Base64 text.") from exc
            signature = hmac.new(secret_bytes, message, hashlib.sha256).digest()
            return {
                "Api-Timestamp": str(timestamp),
                "Api-Key": self.api_key,
                "Api-Signature": base64.b64encode(signature).decode("ascii"),
            }

        if self.token:
            return {self.token_header: self.token}

        raise CamApiError("No CAM API authentication credentials were supplied.")

    def request(self, method: str, path: str, *, body: dict[str, Any] | None = None) -> dict[str, Any]:
        request_path = f"/{path.lstrip('/')}"
        if self.department:
            query = urlencode({"department": self.department}, safe="/")
            request_path = f"{request_path}?{query}"
        url = f"{self.base_url}{request_path}"
        body_text = json.dumps(body) if body is not None else ""
        data = body_text.encode("utf-8") if body is not None else None
        headers = {
            "Accept": "application/json",
            "User-Agent": DEFAULT_USER_AGENT,
            **self.authentication_headers(method, request_path, body_text),
        }
        if data is not None:
            headers["Content-Type"] = "application/json"

        request = Request(url, data=data, headers=headers, method=method)
        try:
            with urlopen(request, timeout=self.timeout) as response:
                payload = json.loads(response.read().decode("utf-8"))
        except HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="replace")
            if exc.code == 401:
                raise CamApiError(
                    "CAM rejected the API credentials or signature. Check the Key/Secret, API "
                    "validity period, IP whitelist, and that this computer's clock is accurate. "
                    f"Server response: {detail}"
                ) from exc
            if exc.code == 403 and "browser_signature_banned" in detail:
                raise CamApiError(
                    "Cloudflare rejected this HTTP client's signature before CAM received the "
                    "request. Confirm that the script is the latest version. "
                    f"Server response: {detail}"
                ) from exc
            raise CamApiError(f"CAM API returned HTTP {exc.code}: {detail}") from exc
        except URLError as exc:
            raise CamApiError(f"Could not connect to CAM API: {exc}") from exc
        except json.JSONDecodeError as exc:
            raise CamApiError("CAM returned a non-JSON response.") from exc

        code = payload.get("code")
        message = payload.get("message")
        if code:
            raise CamApiError(f"CAM API error {code}: {message}")
        return payload

    def list_portfolios(self) -> list[dict[str, Any]]:
        payload = self.request("GET", "/fundv3/openapi/portfolio/list-portfolio")
        return payload.get("result", {}).get("fund_info_list", [])

    def get_asset_positions(
        self,
        fund_names: list[str],
        *,
        snapshot_ns: int,
        quote_source: str,
        volatility_source: str,
    ) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        for start in range(0, len(fund_names), MAX_FUNDS_PER_REQUEST):
            batch = fund_names[start : start + MAX_FUNDS_PER_REQUEST]
            payload = self.request(
                "POST",
                "/anp/openapi/fund-snapshot/get-asset-position",
                body={
                    "fund_list": batch,
                    "start_time": snapshot_ns,
                    "end_time": snapshot_ns,
                    "frequency": "hourly",
                    "quote_source": quote_source,
                    "volatility_source": volatility_source,
                },
            )
            rows.extend(payload.get("result", {}).get("asset_position_overview", []))
        return rows


def filter_portfolios(portfolios: list[dict[str, Any]], tag: str) -> list[dict[str, Any]]:
    target = tag.casefold()
    selected = []
    for portfolio in portfolios:
        aliases = [str(value) for value in portfolio.get("tag_alias_list", [])]
        tag_ids = [str(value) for value in portfolio.get("tag_list", [])]
        if any(value.casefold() == target for value in aliases + tag_ids):
            selected.append(portfolio)
    return selected


def snapshot_datetime(args: argparse.Namespace) -> datetime:
    timezone = ZoneInfo(args.timezone)
    now = datetime.now(timezone)
    date_text = args.date or now.strftime("%Y-%m-%d")
    hour = args.hour if args.hour is not None else now.hour
    return datetime.strptime(f"{date_text} {hour:02d}", "%Y-%m-%d %H").replace(tzinfo=timezone)


def style_table_sheet(sheet, *, freeze_cell: str = "A2") -> None:
    sheet.freeze_panes = freeze_cell
    sheet.auto_filter.ref = sheet.dimensions
    header_fill = PatternFill("solid", fgColor="17365D")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 24

    for column in range(1, sheet.max_column + 1):
        values = [str(sheet.cell(row=row, column=column).value or "") for row in range(1, min(sheet.max_row, 200) + 1)]
        width = min(max(max((len(value) for value in values), default=8) + 2, 11), 38)
        sheet.column_dimensions[get_column_letter(column)].width = width


def add_excel_table(sheet, name: str) -> None:
    if sheet.max_row < 2 or sheet.max_column < 1:
        return
    table = Table(displayName=name, ref=sheet.dimensions)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)


def build_workbook(
    portfolios: list[dict[str, Any]],
    rows: list[dict[str, Any]],
    *,
    tag: str,
    snapshot: datetime,
    output_path: Path,
) -> None:
    portfolio_alias_by_id = {
        str(portfolio.get("fund_name")): str(portfolio.get("fund_alias"))
        for portfolio in portfolios
    }
    for row in rows:
        requested_id = str(row.get("requested_fund_name") or "")
        if not row.get("requested_fund_alias") and requested_id in portfolio_alias_by_id:
            row["requested_fund_alias"] = portfolio_alias_by_id[requested_id]

    rows.sort(
        key=lambda row: (
            str(row.get("requested_fund_alias") or "").casefold(),
            -decimal_or_zero(row.get("usd_exposure_ratio")),
        )
    )

    workbook = Workbook()
    overview = workbook.active
    overview.title = "Overview"
    overview_rows = [
        ["Portfolio Overview Export", ""],
        ["Tag", tag],
        ["Snapshot Time", snapshot.isoformat()],
        ["Portfolio Count", len(portfolios)],
        ["Position Row Count", len(rows)],
        ["Sort", "Portfolio, then USD Exposure % descending"],
    ]
    for row in overview_rows:
        overview.append(row)
    overview["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    overview["A1"].fill = PatternFill("solid", fgColor="17365D")
    overview["B1"].fill = PatternFill("solid", fgColor="17365D")
    overview.column_dimensions["A"].width = 24
    overview.column_dimensions["B"].width = 48

    portfolio_sheet = workbook.create_sheet("Portfolios")
    portfolio_headers = [
        "Portfolio",
        "Portfolio ID",
        "Status",
        "Denomination",
        "Valuation Currency",
        "Tags",
        "Tag IDs",
    ]
    portfolio_sheet.append(portfolio_headers)
    for portfolio in sorted(portfolios, key=lambda item: str(item.get("fund_alias") or "").casefold()):
        portfolio_sheet.append(
            [
                portfolio.get("fund_alias"),
                portfolio.get("fund_name"),
                portfolio.get("status"),
                portfolio.get("denomination"),
                portfolio.get("valuation_currency"),
                ", ".join(portfolio.get("tag_alias_list", [])),
                ", ".join(portfolio.get("tag_list", [])),
            ]
        )
    style_table_sheet(portfolio_sheet)
    add_excel_table(portfolio_sheet, "PortfolioList")

    position_sheet = workbook.create_sheet("Positions")
    position_sheet.append([label for label, _ in SUMMARY_COLUMNS])
    for row in rows:
        values = []
        for _, field in SUMMARY_COLUMNS:
            value = row.get(field)
            if field in {
                "position_size_equity",
                "usd_exposure_ratio",
                "usd_exposure",
                "equity_valuation",
                "mark_price",
                "last_price",
                "upnl_usd",
            }:
                value = float(decimal_or_zero(value))
            values.append(excel_value(value))
        position_sheet.append(values)
    style_table_sheet(position_sheet)
    add_excel_table(position_sheet, "PositionData")

    exposure_column = next(
        index for index, (label, _) in enumerate(SUMMARY_COLUMNS, start=1) if label == "USD Exposure %"
    )
    for cell in position_sheet[get_column_letter(exposure_column)][1:]:
        cell.number_format = "0.00%"

    raw_sheet = workbook.create_sheet("Raw Data")
    raw_headers = sorted({key for row in rows for key in row})
    raw_sheet.append(raw_headers)
    for row in rows:
        raw_sheet.append([excel_value(row.get(header)) for header in raw_headers])
    style_table_sheet(raw_sheet)
    add_excel_table(raw_sheet, "RawPositionData")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export SP Core Asset & Position data from CAM to Excel.")
    parser.add_argument("--base-url", default=DEFAULT_BASE_URL)
    parser.add_argument(
        "--department",
        default=DEFAULT_DEPARTMENT,
        help="CAM department ID. Omit for the API user's default department.",
    )
    parser.add_argument("--tag", default=DEFAULT_TAG)
    parser.add_argument("--timezone", default=DEFAULT_TIMEZONE)
    parser.add_argument("--date", help="Snapshot date in YYYY-MM-DD. Defaults to today in Beijing.")
    parser.add_argument("--hour", type=int, choices=range(24), help="Snapshot hour, 0-23.")
    parser.add_argument("--quote-source", default="cmc_close", choices=("cmc_close", "exchange_first"))
    parser.add_argument(
        "--volatility-source",
        default="exchange",
        choices=("exchange", "linear_interpolation", "svi", "sabr"),
    )
    parser.add_argument("--api-key", help="CAM Open API Key. Prefer CAM_API_KEY or the secure prompt.")
    parser.add_argument(
        "--api-secret",
        help="CAM Open API Secret. Prefer CAM_API_SECRET or the secure prompt.",
    )
    parser.add_argument("--token-header", default="api-token")
    parser.add_argument("--token", help="Optional read-only CAM User Token.")
    parser.add_argument("--output", help="Output .xlsx path.")
    parser.add_argument("--pm-output", help="Top/bottom portfolio workbook output path.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    token = args.token or os.environ.get("CAM_API_TOKEN")
    api_key = args.api_key or os.environ.get("CAM_API_KEY")
    api_secret = args.api_secret or os.environ.get("CAM_API_SECRET")

    if not token and not api_key:
        api_key = getpass.getpass("Enter CAM API Key: ").strip()
    if not token and api_key and not api_secret:
        api_secret = getpass.getpass("Enter CAM API Secret: ").strip()
    if not token and not (api_key and api_secret):
        print("A CAM API Key and Secret are required.", file=sys.stderr)
        return 2

    snapshot = snapshot_datetime(args)
    snapshot_ns = int(snapshot.timestamp() * 1_000_000_000)
    output_path = (
        Path(args.output).expanduser().resolve()
        if args.output
        else OUTPUT_DIR / f"portfolio_overview_{snapshot:%Y%m%d_%H00}.xlsx"
    )
    pm_output_path = (
        Path(args.pm_output).expanduser().resolve()
        if args.pm_output
        else OUTPUT_DIR / f"portfolio_managers_top_bottom_{snapshot:%Y%m%d_%H00}.xlsx"
    )

    client = CamClient(
        base_url=args.base_url,
        department=args.department,
        api_key=api_key,
        api_secret=api_secret,
        token=token,
        token_header=args.token_header,
    )

    try:
        all_portfolios = client.list_portfolios()
        portfolios = filter_portfolios(all_portfolios, args.tag)
        if not portfolios:
            available_tags = sorted(
                {
                    str(tag)
                    for portfolio in all_portfolios
                    for tag in portfolio.get("tag_alias_list", [])
                }
            )
            raise CamApiError(
                f"No portfolio has tag '{args.tag}'. Available tags: {', '.join(available_tags)}"
            )

        fund_names = [str(portfolio["fund_name"]) for portfolio in portfolios]
        print(f"Found {len(fund_names)} portfolio(s) tagged {args.tag}.", flush=True)
        rows = client.get_asset_positions(
            fund_names,
            snapshot_ns=snapshot_ns,
            quote_source=args.quote_source,
            volatility_source=args.volatility_source,
        )
        build_workbook(portfolios, rows, tag=args.tag, snapshot=snapshot, output_path=output_path)
        build_pm_workbook(rows, output_path=pm_output_path)
    except CamApiError as exc:
        print(f"\nExport failed: {exc}", file=sys.stderr)
        return 1

    print(f"Excel export saved to: {output_path}", flush=True)
    print(f"Top/bottom workbook saved to: {pm_output_path}", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
